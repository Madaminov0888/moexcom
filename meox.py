from selenium import webdriver
from selenium.webdriver.support.ui import Select
from bs4 import BeautifulSoup
import datetime
import openpyxl



sdate = datetime.date(year = 2022, month = 2, day = 15)   #boshlanish vaqti
edate = datetime.date(year = 2022, month = 5, day = 15)       # tugash vaqti
month_dct = {1 : 'Января', 2 : 'Февраля', 3:'Марта', 4:'Апреля', 5 :'Мая', 6 : 'Июня', 7: 'Июля', 8 : 'Августа', 9 : 'Сентября', 10 : 'Октября', 11 : 'Ноября', 12 : 'Декабря'}
x = [sdate+datetime.timedelta(days=x) for x in range((edate-sdate).days+1)]


browser = webdriver.Chrome('Desktop/chromedrive')
browser.get('https://www.moex.com/ru/derivatives/open-positions.aspx')


wb = openpyxl.load_workbook('desktop/tgt/moexuchun.xlsx')
sh = wb.active


h = []
s = 4
for j in x:
    s += 1
    element = browser.find_element_by_id("ctl00_PageContent_frmDateTime_CDateDay")
    element1 = browser.find_element_by_id("ctl00_PageContent_frmDateTime_CDateMonth")
    element2 = browser.find_element_by_id("ctl00_PageContent_frmButtom")
    element3 = browser.find_element_by_id("ctl00_PageContent_frmDateTime_CDateYear")
    element.send_keys(f"{j.day}")
    element1.send_keys(month_dct[j.month])
    if j.year != 2022:
        element3.send_keys(f'{j.year}')
    element2.click()
    l = browser.find_element_by_id("redesign-2021")
    html = l.get_attribute('innerHTML')
    soup = BeautifulSoup(html)
    table_tag = soup.find('table', {'class':'table1 _full-width table1'})
    dogovor = table_tag.tbody.find_all('tr')[3].find_all('td')[1:]
    dagavar = table_tag.tbody.find_all('tr')[6].find_all('td')[1:]
    u = []
    for tag_td in dogovor:
        try:
            u.append(tag_td.text.split('\xa0')[0] + tag_td.text.split('\xa0')[1])
        except:
            u.append(tag_td.text)
    u1 = []
    for tg_td in dagavar:
        try:
            u1.append(tg_td.text.split('\xa0')[0] + tg_td.text.split('\xa0')[1])
        except:
            u1.append(tg_td.text)
    
    sh.cell(s, 1).value = f'{j.day}-{j.month}-{j.year}'
    sh.cell(s, 3).value = u[0]
    sh.cell(s, 4).value = u[1]
    sh.cell(s, 5).value = u[2]
    sh.cell(s, 6).value = u[3]
    sh.cell(s, 7).value = u[4]

    sh.cell(s, 9).value = u1[0]
    sh.cell(s, 10).value = u1[1]
    sh.cell(s, 11).value = u1[2]
    sh.cell(s, 12).value = u1[3]
    sh.cell(s, 13).value = u1[4]

wb.save('desktop/tgt/moexuchun.xlsx')

